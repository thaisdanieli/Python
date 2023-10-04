from selenium import webdriver  # Permite fazer as automações
from selenium.webdriver.common.keys import Keys  # Uso do teclado
from selenium.webdriver.common.by import By  # Localizar elementos
from selenium.webdriver.support.select import Select  # Usar DropDown
from time import sleep  # Dar pausas assim que necessário
import openpyxl


numero_OAB = 133864

# 1.1 Entrar no site - https://pje-consulta-publica.tjmg.jus.br/
driver = webdriver.Chrome()
driver.get('https://pje-consulta-publica.tjmg.jus.br/')
sleep(30)

# 1.2 Digitar numero oab
campo_OAB = driver.find_element(
    By.XPATH, "//input[@id='fPP:Decoration:numeroOAB']")
campo_OAB.send_keys(numero_OAB)

# selecionar estado
dropdown_estados = driver.find_element(
    By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']")
opcoes_estados = Select(dropdown_estados)
opcoes_estados.select_by_visible_text('SP')

# 1.3 clicar em pesquisar
botao_pesquisar = driver.find_element(
    By.XPATH, "//input[@id='fPP:searchProcessos']")
botao_pesquisar.click()
sleep(10)
# 1.4 entrar em cada um dos processos

processos = driver.find_elements(By.XPATH, "//b[@class='btn-block']") # Vamos trabalhar com vários elementos

for processo in processos:
    processo.click()
    sleep(10)
    
    janelas = driver.window_handles # Irá verificar quais janelas/paginas estão disponiveis
    
    driver.switch_to.window(janelas[-1]) # Enviar novos comandos para essa nova janela
    
    driver.set_window_size(1920, 1080) # Definir tamanho fixo da janela/resolução fixa (tamanho)

    # 1.5 extrair o nº do processo e data da distribuição
    numero_processo = driver.find_elements(
        By.XPATH, "//div[@class='col-sm-12 ']")
    numero_processo = numero_processo[0]  # indice 1
    numero_processo = numero_processo.text  # serve para guardar apenas o texto

    data_distribuicao =  driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")
    data_distribuicao = data_distribuicao[1]
    data_distribuicao = data_distribuicao.text


# 1.6 extrair e guardar todas as últimas movimentações
movimentacoes =  driver.find_elements(By.XPATH, "//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class='rich-table-row')]//td//div//div//span")

lista_movimentacoes = []
for movimentacao in movimentacoes:     
    lista_movimentacoes.append(movimentacao.text) # Adicionar movimentacao que está sendo interada no momento



# 1.7 guardar tudo no excel, separados por processo
workbook = openpyxl.load_workbook('dados.xlsm') # Carregar a planilha

try:
    #código para inserir dados em pagina já existente
    # acessar página do process
    pagina_processo = workbook[numero_processo]

    # Criar nome das colunas
    pagina_processo['A1'].value = "Número Processo"
    pagina_processo['B1'].value = "Data Distribuição"
    pagina_processo['C1'].value = "Movimentações"

    # Adicionar número do processo
    pagina_processo['A2'].value = numero_processo

    # Adicionar data de distribuição
    pagina_processo['B2'].value = data_distribuicao

    # Adicionar movimentações
    for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes), 
                                                            min_col=3, max_col=3)):
    # Especificamos a localização de onde as informações

    # serão adicionadas: Número da linha 2, coluna 3
    # MAX_ROX=LEN > Máximo de linhas/Len retorna o tamanho da lista
        for celula in linha:
            celula.value = lista_movimentacoes[index]
    workbook.save('dados.xlsx')
    driver.close()
    driver.switch_to.window(driver.window_handles[0])

    sleep(5)
except Exception as error:
    #Código para criar uma página do zero e inserir as informações
    workbook.create_sheet(numero_processo)

    pagina_processo = workbook[numero_processo]

    # Criar nome das colunas
    pagina_processo['A1'].value = "Número Processo"
    pagina_processo['B1'].value = "Data Distribuição"
    pagina_processo['C1'].value = "Movimentações"

    # Adicionar número do processo
    pagina_processo['A2'].value = numero_processo

    # Adicionar data de distribuição
    pagina_processo['B2'].value = data_distribuicao

    # Adicionar movimentações
    for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes),min_col=3, max_col=3)): 
    # Especificamos a localização de onde as informações
    # serão adicionadas: Número da linha 2, coluna 3
    # MAX_ROX=LEN > Máximo de linhas/Len retorna o tamanho da lista
        for celula in linha:
            celula.value = lista_movimentacoes[index]
            
    workbook.save('dados.xlsx')
    driver.close()
    driver.switch_to.window(driver.window_handles[0])