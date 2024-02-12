from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By  
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select  
from time import sleep  
import openpyxl
import re

def caracterEspecial(frase):
    frase = ''.join(c for c in frase if c.isalnum() or c.isspace())
    return frase

driver = webdriver.Chrome()
driver.get('https://www.themoviedb.org/')
sleep(5)

acept = driver.find_element(By.XPATH, '//button[@id="onetrust-accept-btn-handler"]')
acept.click()
sleep(1)

movies = driver.find_elements(By.XPATH, "//div[@class='card style_1']")

for movie in movies:    
    # Use ActionChains para realizar a combinação de teclas Ctrl (ou Command) + clique
    actions = ActionChains(driver)
    actions.key_down(Keys.CONTROL).click(movie).key_up(Keys.CONTROL).perform()

    sleep(2)

    # Alterne para a nova aba
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])

    try:
        name_Movie  = driver.find_element(By.XPATH, "//div[@class='title ott_false']//a")
    except:
        name_Movie  = driver.find_element(By.XPATH, "//div[@class='title ott_true']//a")

    name_Movie = name_Movie.text

    name_Movie = caracterEspecial(name_Movie)

    type_Movie      = driver.find_element(By.XPATH, "//div[@class='facts']//span[@class='genres']")
    type_Movie      = type_Movie.text

    sinop           = driver.find_element(By.XPATH, "//div[@class='header_info']//div[@class='overview']")
    sinop           = sinop.text

    workbook = openpyxl.load_workbook('TheMoviesDB_Python.xlsx')   

    try:
        workbook.create_sheet(name_Movie)
        pagina_excel = workbook[name_Movie]

        pagina_excel['A1'].value = "Filme"
        pagina_excel['B1'].value = "Gênero"
        pagina_excel['C1'].value = "Sinopse"

        pagina_excel['A2'].value = name_Movie
        pagina_excel['B2'].value = type_Movie
        pagina_excel['C2'].value = sinop

        workbook.save('TheMoviesDB_Python.xlsx')
        driver.back()
        driver.refresh()

    except Exception as error:
        workbook.create_sheet(name_Movie)
        pagina_excel = workbook[name_Movie]      

        pagina_excel['A1'].value = "Filme"
        pagina_excel['B1'].value = "Gênero"
        pagina_excel['C1'].value = "Sinopse"

        pagina_excel['A2'].value = name_Movie
        pagina_excel['B2'].value = type_Movie
        pagina_excel['C2'].value = sinop

        workbook.save('TheMoviesDB_Python.xlsx') 
        driver.back()
        driver.refresh()

    driver.close()

    # Volte para a aba original (opcional)
    driver.switch_to.window(janelas[0])
